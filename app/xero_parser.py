"""
Xero-specific Excel parser for CFO dashboard.

Xero exports have a consistent 4-row header:
  Row 1: Company name
  Row 2: Report type ("Profit and Loss", "Balance Sheet", "Trial Balance")
  Row 3: Date range
  Row 4: Blank
  Row 5+: Data

Functions:
  is_xero_format(path) -> bool
  get_xero_report_type(path) -> str | None
  parse_xero_pl(path) -> dict
  parse_xero_balance_sheet(path) -> dict
  parse_xero_trial_balance(path) -> dict
"""

import re
import openpyxl
from pathlib import Path

# Known Xero report type strings → internal keys
_REPORT_TYPE_MAP = {
    "profit and loss":  "pl",
    "balance sheet":    "balance_sheet",
    "trial balance":    "trial_balance",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_ws(path: Path):
    """Load the first worksheet from an Excel file."""
    wb = openpyxl.load_workbook(Path(path), data_only=True)
    return wb[wb.sheetnames[0]]


def _cell_str(ws, row: int, col: int = 1) -> str:
    """Return cell value as a stripped string, or '' if empty/None."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def is_xero_format(path) -> bool:
    """
    Return True if the Excel file looks like a Xero export.

    Criteria:
      - Row 1, col A: non-empty string (company name)
      - Row 2, col A: recognised Xero report type
      - Row 4: completely blank (all cells in first 15 cols are None/empty)
    """
    try:
        ws = _load_ws(path)
        row1 = _cell_str(ws, 1)
        row2 = _cell_str(ws, 2).lower()
        # Row 4 must be blank
        row4_values = [ws.cell(row=4, column=c).value for c in range(1, 16)]
        row4_blank = all(v is None or str(v).strip() == "" for v in row4_values)

        return bool(row1) and (row2 in _REPORT_TYPE_MAP) and row4_blank
    except Exception:
        return False


def get_xero_report_type(path) -> "str | None":
    """
    Return 'pl', 'balance_sheet', 'trial_balance', or None.
    """
    try:
        ws = _load_ws(path)
        row2 = _cell_str(ws, 2).lower()
        return _REPORT_TYPE_MAP.get(row2)
    except Exception:
        return None


def _parse_amount(v) -> "float | None":
    """
    Parse a cell value into a float.

    Handles:
      None          → None
      int/float     → float
      "(123,456)"   → -123456.0   (parentheses = negative)
      "-"           → 0.0
      ""            → None
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s == "-":
        return 0.0 if s == "-" else None
    # Parentheses notation: (123,456) → -123456
    if s.startswith("(") and s.endswith(")"):
        inner = s[1:-1].replace(",", "")
        try:
            return -abs(float(inner))
        except ValueError:
            return None
    # Plain string number with commas
    s_clean = s.replace(",", "")
    try:
        return float(s_clean)
    except ValueError:
        return None


def _read_data_rows(path) -> list:
    """
    Read rows 5+ from the first sheet as a list of value tuples (up to 15 cols).
    """
    ws = _load_ws(path)
    rows = []
    max_row = ws.max_row or 0
    for r in range(5, max_row + 1):
        row_vals = tuple(ws.cell(row=r, column=c).value for c in range(1, 16))
        rows.append(row_vals)
    return rows


def _find_amount_col(rows: list) -> int:
    """
    Find which column (1-based index into tuple) holds the primary amounts.

    Strategy: scan first 15 rows; find the first column index >= 1 that
    has a numeric value in a row where col 0 (label) is non-empty.
    Default: 1.
    """
    for row in rows[:15]:
        label = row[0]
        if label is None or str(label).strip() == "":
            continue
        for col_idx in range(1, min(15, len(row))):
            v = row[col_idx]
            if v is not None and isinstance(v, (int, float)):
                return col_idx
            if v is not None and isinstance(v, str):
                parsed = _parse_amount(v)
                if parsed is not None:
                    return col_idx
    return 1


# ── P&L Parser ────────────────────────────────────────────────────────────────

def parse_xero_pl(path) -> dict:
    """
    Parse a Xero Profit and Loss Excel export.

    Returns a dict with keys matching the Claude extractor output:
      Total Revenue, Gross Profit, EBITDA, Net Profit After Tax,
      Food & Beverage Costs, Payroll & Staff Costs, Rent & Utilities,
      Marketing & Advertising
    """
    rows = _read_data_rows(path)
    amt_col = _find_amount_col(rows)

    result = {
        "Total Revenue":         None,
        "Gross Profit":          None,
        "EBITDA":                None,
        "Net Profit After Tax":  None,
        "Food & Beverage Costs": None,
        "Payroll & Staff Costs": None,
        "Rent & Utilities":      None,
        "Marketing & Advertising": None,
    }

    # Accumulators for individual lines (used only if no Total row found)
    _payroll_acc   = 0.0
    _rent_acc      = 0.0
    _marketing_acc = 0.0
    dna_total      = 0.0
    interest_total = 0.0
    tax_total      = 0.0

    # Flags: set True once a "Total" row for that category is found
    _payroll_total_found   = False
    _rent_total_found      = False
    _marketing_total_found = False

    for row in rows:
        label_raw = row[0]
        if label_raw is None:
            continue
        label = str(label_raw).strip()
        if not label:
            continue

        label_lower = label.lower()
        amt = _parse_amount(row[amt_col] if amt_col < len(row) else None)

        # ── Revenue ─────────────────────────────────────────────────────────
        if label_lower in (
            "total income", "total revenue", "total operating revenue",
            "total trading income",
        ):
            result["Total Revenue"] = amt
            continue

        # ── Cost of Sales / Food & Beverage ─────────────────────────────────
        if label_lower in ("total cost of sales", "total cogs"):
            result["Food & Beverage Costs"] = abs(amt) if amt is not None else None
            continue

        # ── Gross Profit ─────────────────────────────────────────────────────
        if label_lower in ("gross profit", "gross profit/(loss)"):
            result["Gross Profit"] = amt
            continue

        # ── Net Profit ───────────────────────────────────────────────────────
        if label_lower in (
            "net profit", "net profit/(loss)", "profit for the year", "net income",
        ):
            result["Net Profit After Tax"] = amt
            continue

        # ── Payroll ──────────────────────────────────────────────────────────
        payroll_match = any(kw in label_lower for kw in
                            ("payroll", "wages", "salaries", "staff costs"))
        if payroll_match:
            if label.startswith("Total") or label_lower.startswith("total"):
                result["Payroll & Staff Costs"] = amt
                _payroll_total_found = True
            elif not _payroll_total_found and amt is not None:
                _payroll_acc += abs(amt)
            continue

        # ── Rent & Utilities ─────────────────────────────────────────────────
        rent_match = any(kw in label_lower for kw in ("rent", "utilities", "occupancy"))
        if rent_match:
            if label.startswith("Total") or label_lower.startswith("total"):
                result["Rent & Utilities"] = amt
                _rent_total_found = True
            elif not _rent_total_found and amt is not None:
                _rent_acc += abs(amt)
            continue

        # ── Marketing & Advertising ──────────────────────────────────────────
        marketing_match = any(kw in label_lower for kw in ("marketing", "advertising"))
        if marketing_match:
            if label.startswith("Total") or label_lower.startswith("total"):
                result["Marketing & Advertising"] = amt
                _marketing_total_found = True
            elif not _marketing_total_found and amt is not None:
                _marketing_acc += abs(amt)
            continue

        # ── D&A ──────────────────────────────────────────────────────────────
        if any(kw in label_lower for kw in
               ("depreciation", "amortisation", "amortization")):
            if amt is not None:
                dna_total += abs(amt)
            continue

        # ── Interest / Finance ───────────────────────────────────────────────
        if any(kw in label_lower for kw in
               ("interest expense", "finance cost", "bank charges")):
            if amt is not None:
                interest_total += abs(amt)
            continue

        # ── Tax ──────────────────────────────────────────────────────────────
        if any(kw in label_lower for kw in ("income tax", "tax expense")):
            if amt is not None:
                tax_total += abs(amt)
            continue

    # Fill accumulators where no Total row was found
    if not _payroll_total_found and _payroll_acc:
        result["Payroll & Staff Costs"] = _payroll_acc
    if not _rent_total_found and _rent_acc:
        result["Rent & Utilities"] = _rent_acc
    if not _marketing_total_found and _marketing_acc:
        result["Marketing & Advertising"] = _marketing_acc

    # Compute EBITDA
    net = result.get("Net Profit After Tax")
    if net is not None:
        result["EBITDA"] = net + dna_total + interest_total + tax_total

    return result


# ── Balance Sheet Parser ──────────────────────────────────────────────────────

def parse_xero_balance_sheet(path) -> dict:
    """
    Parse a Xero Balance Sheet Excel export.

    Returns a dict with keys matching the Claude extractor output.
    """
    rows = _read_data_rows(path)
    amt_col = _find_amount_col(rows)

    result = {
        "TOTAL ASSETS":                   None,
        "Total Current Assets":            None,
        "Total Non-Current Assets":        None,
        "Total Current Liabilities":       None,
        "Total Non-Current Liabilities":   None,
        "Total Equity":                    None,
        "Cash & Cash Equivalents":         None,
        "Accounts Receivable":             None,
        "Inventory":                       None,
        "Prepaid Expenses":                None,
        "Net PPE":                         None,
        "Intangible Assets":               None,
    }

    # Accumulators for individual lines
    _cash_acc      = 0.0
    _ar_acc        = 0.0
    _inv_acc       = 0.0
    _prepaid_acc   = 0.0
    _ppe_acc       = 0.0
    _intangible_acc = 0.0

    _cash_total_found      = False
    _ar_total_found        = False
    _inv_total_found       = False
    _prepaid_total_found   = False
    _ppe_total_found       = False
    _intangible_total_found = False

    for row in rows:
        label_raw = row[0]
        if label_raw is None:
            continue
        label = str(label_raw).strip()
        if not label:
            continue

        label_lower = label.lower()
        amt = _parse_amount(row[amt_col] if amt_col < len(row) else None)
        is_total = label.startswith("Total") or label_lower.startswith("total")

        # ── Total rows ───────────────────────────────────────────────────────

        if label_lower in ("total current assets", "total current"):
            result["Total Current Assets"] = amt
            continue

        if label_lower in (
            "total non-current assets", "total fixed assets",
            "total non current assets",
        ):
            result["Total Non-Current Assets"] = amt
            continue

        if label_lower == "total assets":
            result["TOTAL ASSETS"] = amt
            continue

        if label_lower == "total current liabilities":
            result["Total Current Liabilities"] = amt
            continue

        if label_lower in (
            "total non-current liabilities", "total long-term liabilities",
            "total non current liabilities",
        ):
            result["Total Non-Current Liabilities"] = amt
            continue

        if label_lower in (
            "total equity", "total shareholders' equity",
            "total shareholders equity", "net assets",
        ):
            result["Total Equity"] = amt
            continue

        # ── Individual asset/liability lines ─────────────────────────────────

        # Cash
        if any(kw in label_lower for kw in
               ("cash at bank", "bank account", "cheque", "cash and cash")):
            if amt is not None:
                _cash_acc += amt  # can be negative (overdraft)
            continue

        # Accounts Receivable
        if any(kw in label_lower for kw in
               ("accounts receivable", "trade receivable", "debtors")):
            if amt is not None:
                _ar_acc += abs(amt)
            continue

        # Inventory
        if any(kw in label_lower for kw in ("inventory", "stock on hand")):
            if amt is not None:
                _inv_acc += abs(amt)
            continue

        # Prepaid
        if any(kw in label_lower for kw in ("prepaid", "prepayment")):
            if amt is not None:
                _prepaid_acc += abs(amt)
            continue

        # Net PPE
        if label_lower == "net ppe" or (
            is_total and any(kw in label_lower for kw in
                             ("property, plant", "fixed asset", "plant and equip"))
        ):
            result["Net PPE"] = amt
            _ppe_total_found = True
            continue

        # PPE components (accumulate, net of depreciation via signed amounts)
        if any(kw in label_lower for kw in
               ("property, plant", "plant and equip", "equipment",
                "accumulated depreciation", "accumulated amortisation")):
            if not _ppe_total_found and amt is not None:
                _ppe_acc += amt  # depreciation stored as negative number
            continue

        # Intangibles
        if "intangible" in label_lower:
            if is_total:
                result["Intangible Assets"] = amt
                _intangible_total_found = True
            elif not _intangible_total_found and amt is not None:
                _intangible_acc += abs(amt)
            continue

    # Fill individual accumulators
    if _cash_acc != 0.0:
        result["Cash & Cash Equivalents"] = _cash_acc
    if _ar_acc:
        result["Accounts Receivable"] = _ar_acc
    if _inv_acc:
        result["Inventory"] = _inv_acc
    if _prepaid_acc:
        result["Prepaid Expenses"] = _prepaid_acc
    if not _ppe_total_found and _ppe_acc != 0.0:
        result["Net PPE"] = _ppe_acc
    if not _intangible_total_found and _intangible_acc:
        result["Intangible Assets"] = _intangible_acc

    # Derive TOTAL ASSETS if missing
    if result["TOTAL ASSETS"] is None:
        ca  = result.get("Total Current Assets")     or 0
        nca = result.get("Total Non-Current Assets") or 0
        if ca or nca:
            result["TOTAL ASSETS"] = ca + nca

    return result


# ── Trial Balance Parser ──────────────────────────────────────────────────────

def parse_xero_trial_balance(path) -> dict:
    """
    Parse a Xero Trial Balance Excel export.

    Returns {"accounts": [{"name": str, "debit": float, "credit": float}]}.
    Skips header/total/blank rows and rows where both debit and credit are 0.
    """
    rows = _read_data_rows(path)

    skip_labels = {"", "account", "total", "totals"}
    accounts = []

    for row in rows:
        label_raw = row[0]
        if label_raw is None:
            continue
        label = str(label_raw).strip()
        if label.lower() in skip_labels:
            continue

        # Debit in col 1, credit in col 2 (Xero TB layout)
        debit  = _parse_amount(row[1] if len(row) > 1 else None) or 0.0
        credit = _parse_amount(row[2] if len(row) > 2 else None) or 0.0

        if debit == 0.0 and credit == 0.0:
            continue

        accounts.append({"name": label, "debit": debit, "credit": credit})

    return {"accounts": accounts}
