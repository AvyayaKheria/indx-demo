"""
AI-powered financial data extraction using the Claude API.

Reads Excel files in any real-world format (different layouts, merged cells,
extra header rows, varied column names) and returns clean, normalised JSON
that the dashboard can consume.

Usage:
    from .extractor import extract_all, ExtractionError
    summary = extract_all(session_dir)   # raises ExtractionError on failure
"""

import os
import json
import openpyxl
import anthropic
from pathlib import Path

MODEL = "claude-sonnet-4-20250514"
MAX_CHARS = 14_000   # ~3 500 tokens — plenty for financial tables


# ── Custom exception ──────────────────────────────────────────────────────────

class ExtractionError(ValueError):
    """Raised when Claude cannot extract meaningful data from a file."""
    pass


# ── Helpers ───────────────────────────────────────────────────────────────────

def _client() -> anthropic.Anthropic:
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        raise ExtractionError(
            "ANTHROPIC_API_KEY is not configured. "
            "Add it as an environment variable to enable AI-powered extraction."
        )
    return anthropic.Anthropic(api_key=key)


def _xl_to_text(path: Path) -> str:
    """Convert every sheet of an Excel file to a tab-separated text block."""
    wb = openpyxl.load_workbook(path, data_only=True)
    lines: list[str] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        lines.append(f"=== Sheet: {sname} ===")
        for row in ws.iter_rows(max_row=150, max_col=20, values_only=True):
            if any(v is not None for v in row):
                lines.append("\t".join(
                    "" if v is None else str(v).strip()
                    for v in row
                ))
    text = "\n".join(lines)
    return text[:MAX_CHARS]


def _call(prompt: str) -> dict:
    """Send a prompt to Claude and parse the JSON response."""
    resp = _client().messages.create(
        model=MODEL,
        max_tokens=2048,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = resp.content[0].text.strip()

    # Strip any markdown code fences Claude might add
    for fence in ("```json", "```"):
        if fence in raw:
            raw = raw.split(fence, 1)[-1].rsplit("```", 1)[0].strip()
            break

    # Find the outermost JSON object
    start = raw.find("{")
    if start != -1:
        raw = raw[start:]

    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ExtractionError(
            f"Claude returned a response that could not be parsed as JSON.\n"
            f"Raw response (first 400 chars): {raw[:400]}"
        ) from exc


# ── Per-file extractors ───────────────────────────────────────────────────────

def _extract_revenue(path: Path) -> dict:
    xl = _xl_to_text(path)
    prompt = f"""Extract monthly revenue data from this financial spreadsheet.

Return ONLY a valid JSON object — no explanation, no markdown fences:

{{
  "months": ["Jan", "Feb", ...],
  "channels": {{
    "Channel Name": [number, ...]
  }},
  "totals": [number, ...]
}}

Rules:
- months: ordered list of month labels — use short names like "Jan", "Feb" etc.
- channels: one key per revenue stream/channel; value is an array of monthly amounts
  aligned positionally to months (same length)
- totals: monthly grand-total revenue amounts aligned to months
- All amounts are plain numbers — no $ signs, commas, currency text
- Do NOT include TOTAL or annual summary rows — monthly rows only
- If the file has multiple sheets, use the one with monthly revenue detail
- Use whatever channel names appear (e.g. "Dine-In", "Delivery", "Online", "Walk-in")

Spreadsheet content:
{xl}"""

    data = _call(prompt)
    if not isinstance(data.get("months"), list) or not isinstance(data.get("channels"), dict):
        raise ExtractionError(
            "Revenue Sheet: could not find monthly revenue by channel.\n"
            "Expected: month rows with revenue broken down by category (e.g. Dine-In, Delivery)."
        )
    if not data["channels"]:
        raise ExtractionError(
            "Revenue Sheet: no revenue channels found. "
            "Please check the file contains revenue data broken down by channel."
        )
    return data


def _extract_costs(path: Path) -> dict:
    xl = _xl_to_text(path)
    prompt = f"""Extract monthly cost data from this financial spreadsheet.

Return ONLY a valid JSON object:

{{
  "months": ["Jan", "Feb", ...],
  "categories": {{
    "Category Name": [number, ...]
  }},
  "totals": [number, ...]
}}

Rules:
- months: ordered month labels
- categories: one key per cost type; value is monthly amounts aligned to months
- totals: monthly grand-total cost amounts
- All amounts are plain numbers
- Do NOT include TOTAL rows — monthly rows only
- Common cost categories: COGS, Food Costs, Payroll, Staff, Rent, Utilities, Marketing, Admin

Spreadsheet content:
{xl}"""

    data = _call(prompt)
    if not isinstance(data.get("months"), list) or not isinstance(data.get("categories"), dict):
        raise ExtractionError(
            "Cost Sheet: could not find monthly costs by category.\n"
            "Expected: month rows with costs broken down by type (e.g. COGS, Payroll, Rent)."
        )
    if not data["categories"]:
        raise ExtractionError(
            "Cost Sheet: no cost categories found. "
            "Please check the file contains monthly cost data."
        )
    return data


def _extract_pl(path: Path) -> dict:
    # Try Xero parser first (fast, no API call required)
    from .xero_parser import is_xero_format, get_xero_report_type, parse_xero_pl
    if is_xero_format(path) and get_xero_report_type(path) == "pl":
        result = parse_xero_pl(path)
        if result.get("Total Revenue"):
            return result

    xl = _xl_to_text(path)
    prompt = f"""Extract P&L (income statement) data from this financial spreadsheet.

Map line items to these EXACT key names where a match exists:
- "Total Revenue"          → total sales / net revenue / turnover / total income
- "Gross Profit"           → revenue minus cost of goods sold
- "EBITDA"                 → earnings before interest, tax, depreciation & amortisation
- "Net Profit After Tax"   → bottom-line profit / net income / profit for the year
- "Food & Beverage Costs"  → COGS / food costs / cost of sales
- "Payroll & Staff Costs"  → wages / salaries / staff costs / labour
- "Rent & Utilities"       → rent / lease / occupancy / utilities
- "Marketing & Advertising"→ marketing / advertising / promotions

Return ONLY a valid JSON object using those exact key names:

{{
  "Total Revenue": number,
  "Gross Profit": number or null,
  "EBITDA": number or null,
  "Net Profit After Tax": number or null,
  "Food & Beverage Costs": number or null,
  "Payroll & Staff Costs": number or null,
  "Rent & Utilities": number or null,
  "Marketing & Advertising": number or null
}}

Rules:
- Values are plain numbers (positive for income/profit, negative for losses/expenses)
- Use null for any item that cannot be found
- Use full-year totals for annual P&L statements

Spreadsheet content:
{xl}"""

    data = _call(prompt)
    if not data.get("Total Revenue"):
        raise ExtractionError(
            "P&L Statement: could not find Total Revenue.\n"
            "Please check the file contains an income statement or P&L."
        )
    return data


def _extract_balance_sheet(path: Path) -> dict:
    # Try Xero parser first (fast, no API call required)
    from .xero_parser import is_xero_format, get_xero_report_type, parse_xero_balance_sheet
    if is_xero_format(path) and get_xero_report_type(path) == "balance_sheet":
        result = parse_xero_balance_sheet(path)
        if result.get("TOTAL ASSETS"):
            return result

    xl = _xl_to_text(path)
    prompt = f"""Extract balance sheet data from this financial spreadsheet.

Return ONLY a valid JSON object using these EXACT key names:

{{
  "TOTAL ASSETS": number or null,
  "Total Current Assets": number or null,
  "Total Non-Current Assets": number or null,
  "Total Current Liabilities": number or null,
  "Total Non-Current Liabilities": number or null,
  "Total Equity": number or null,
  "Cash & Cash Equivalents": number or null,
  "Accounts Receivable": number or null,
  "Inventory": number or null,
  "Prepaid Expenses": number or null,
  "Net PPE": number or null,
  "Intangible Assets": number or null
}}

Mapping guidance:
- "Total Current Assets"          = cash + receivables + inventory + prepaid (short-term)
- "Total Non-Current Assets"      = property, equipment, intangibles (long-term)
- "TOTAL ASSETS"                  = all assets combined
- "Total Current Liabilities"     = short-term obligations due within 1 year
- "Total Non-Current Liabilities" = long-term debt, non-current obligations
- "Total Equity"                  = shareholders equity / owners equity / net assets
- "Net PPE"                       = property, plant & equipment net of depreciation

Rules:
- All values are plain numbers; use null if not found
- If TOTAL ASSETS is missing, sum current + non-current assets

Spreadsheet content:
{xl}"""

    data = _call(prompt)
    # Compute TOTAL ASSETS from components if missing
    if not data.get("TOTAL ASSETS"):
        ca  = data.get("Total Current Assets")     or 0
        nca = data.get("Total Non-Current Assets") or 0
        if ca or nca:
            data["TOTAL ASSETS"] = ca + nca

    if not data.get("TOTAL ASSETS"):
        raise ExtractionError(
            "Balance Sheet: could not find asset totals.\n"
            "Please check the file contains a balance sheet."
        )
    return data


def _extract_trial_balance(path: Path) -> dict:
    # Try Xero parser first (fast, no API call required)
    from .xero_parser import is_xero_format, get_xero_report_type, parse_xero_trial_balance
    if is_xero_format(path) and get_xero_report_type(path) == "trial_balance":
        result = parse_xero_trial_balance(path)
        if result.get("accounts"):
            return result

    xl = _xl_to_text(path)
    prompt = f"""Extract trial balance data from this financial spreadsheet.

Return ONLY a valid JSON object:

{{
  "accounts": [
    {{"name": "Account Name", "debit": number, "credit": number}},
    ...
  ]
}}

Rules:
- Include all accounts with debit and credit balances
- Use 0 (not null) when a column is blank
- All values are plain numbers
- Exclude TOTAL rows, header rows, and blank rows

Spreadsheet content:
{xl}"""

    data = _call(prompt)
    if not isinstance(data.get("accounts"), list):
        raise ExtractionError(
            "Trial Balance: could not find account data.\n"
            "Please check the file contains a trial balance."
        )
    return data


# ── Public entry point ────────────────────────────────────────────────────────

def extract_all(session_dir: Path) -> dict:
    """
    Run AI extraction for all 5 uploaded files in session_dir.

    Saves <name>_extracted.json and extraction_summary.json to session_dir.
    Returns the summary dict shown on the confirmation screen.
    Raises ExtractionError with a user-friendly message on any failure.
    """
    session_dir = Path(session_dir)

    jobs = [
        ("revenue.xlsx",       _extract_revenue,       "revenue_extracted.json",       "Revenue Sheet"),
        ("costs.xlsx",         _extract_costs,         "costs_extracted.json",         "Cost Sheet"),
        ("pl.xlsx",            _extract_pl,            "pl_extracted.json",            "P&L Statement"),
        ("balance_sheet.xlsx", _extract_balance_sheet, "balance_sheet_extracted.json", "Balance Sheet"),
        ("trial_balance.xlsx", _extract_trial_balance, "trial_balance_extracted.json", "Trial Balance"),
    ]

    extracted: dict[str, dict] = {}
    for xlsx_name, fn, json_name, label in jobs:
        try:
            data = fn(session_dir / xlsx_name)
            (session_dir / json_name).write_text(json.dumps(data, indent=2))
            key = json_name.replace("_extracted.json", "")
            extracted[key] = data
        except ExtractionError:
            raise
        except Exception as exc:
            raise ExtractionError(
                f"{label}: unexpected error during extraction — {exc}"
            ) from exc

    # Build summary for confirmation screen
    pl_data   = extracted.get("pl",            {})
    bs_data   = extracted.get("balance_sheet", {})
    rev_data  = extracted.get("revenue",       {})
    cost_data = extracted.get("costs",         {})

    def _n(v):
        return v if isinstance(v, (int, float)) else None

    summary = {
        "total_revenue":    _n(pl_data.get("Total Revenue")),
        "gross_profit":     _n(pl_data.get("Gross Profit")),
        "ebitda":           _n(pl_data.get("EBITDA")),
        "net_profit":       _n(pl_data.get("Net Profit After Tax")),
        "total_assets":     _n(bs_data.get("TOTAL ASSETS")),
        "total_equity":     _n(bs_data.get("Total Equity")),
        "revenue_channels": list(rev_data.get("channels",   {}).keys()),
        "cost_categories":  list(cost_data.get("categories", {}).keys()),
        "monthly_rows":     len(rev_data.get("months", [])),
    }

    (session_dir / "extraction_summary.json").write_text(
        json.dumps(summary, indent=2)
    )
    return summary
